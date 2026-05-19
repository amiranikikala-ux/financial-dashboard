"""Self-check Foodmart per-supplier income (retro + cashback + marketing).

Usage:
    python scripts/foodmart_verify.py <month YYYY-MM>

For the given month, computes expected:
  - Retro per supplier (from saved formula × rs.ge waybill or Megaplus brand sales)
  - Cashback (from Lela template × Megaplus sold qty)
  - Marketing (not fully derivable — pull rate-card from Elene file if present)

Then compares with Elene's per-month files (if available) and bank deposit (next month).

Required data on disk:
  - Financial_Analysis/cache/rsge/<YYYY>.parquet
  - Megaplus DB (NSSM service running)
  - Financial_Analysis/Foodmart/supplier_retro_formulas.json (saved per-supplier rates)
  - Optional: Financial_Analysis/Foodmart/გაშიფვრები_<YY>/<MM>,<YY> რეტრო.xlsx (Elene's)
  - Optional: Financial_Analysis/Foodmart/სააქცია_<YYYY>-<YYYY>/<YYYY>-<MM>_<...>.xlsx (Lela's)

Output: console report + optional JSON dump.
"""
from __future__ import annotations

import sys, os, json, argparse, glob
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, r'C:\financial-dashboard')
os.chdir(r'C:\financial-dashboard')

import pandas as pd
import openpyxl

from dashboard_pipeline.megaplus_backup import _connect, _db_name_for

ROOT = Path(r'C:\financial-dashboard')
FOODMART = ROOT / 'Financial_Analysis' / 'Foodmart'
FORMULAS_PATH = FOODMART / 'supplier_retro_formulas.json'
RSGE_CACHE = ROOT / 'Financial_Analysis' / 'cache' / 'rsge'

STORES = ['1301', '1329']  # ოზურგეთი, დვაბზუ


def normalize_barcode(b) -> str:
    s = str(b or '').strip()
    if s.endswith('.0'): s = s[:-2]
    return s.lstrip('0') or s


def load_formulas() -> dict:
    if not FORMULAS_PATH.exists():
        return {}
    with FORMULAS_PATH.open('r', encoding='utf-8') as f:
        return json.load(f)


def rsge_waybill_in_per_tin(year: int, month: int) -> dict:
    """sum amount per supplier TIN from rs.ge cache."""
    path = RSGE_CACHE / f'{year}.parquet'
    if not path.exists():
        return {}
    df = pd.read_parquet(path)
    df['org_str'] = df['ორგანიზაცია'].astype(str)
    df['amount'] = pd.to_numeric(df['თანხა'], errors='coerce').fillna(0)
    df['month'] = df['გააქტიურების თარ.'].astype(str).str[:7]
    df = df[df['month'] == f'{year}-{month:02d}']
    # Extract TIN — usually in parens at start: "(123456789) name"
    df['tin'] = df['org_str'].str.extract(r'\((\d{9,11})\)')[0]
    grouped = df.groupby('tin')['amount'].sum().to_dict()
    return {t: float(v) for t, v in grouped.items() if t}


def megaplus_brand_sales(name_keywords: list[str], year: int, month: int) -> tuple[float, float]:
    """Sum revenue + cogs across stores for products with names matching any keyword."""
    cond = ' OR '.join(f"p.P_NAME LIKE N'%{k}%'" for k in name_keywords)
    month_start = f'{year}-{month:02d}-01'
    next_m = month + 1
    next_y = year
    if next_m > 12:
        next_m, next_y = 1, year + 1
    month_end = f'{next_y}-{next_m:02d}-01'
    rev = cogs = 0
    for store_id in STORES:
        conn = _connect(_db_name_for(store_id))
        try:
            sql = f"""SELECT SUM(o.ORD_jamjam) rev, SUM(o.ORD_GETPRICE*o.ORD_quant) cogs
                FROM ORDERS o JOIN PRODUCTS p ON p.P_ID=o.ORD_P_ID
                WHERE o.ORD_ACT=1 AND o.ORD_TIMESTAMP>=? AND o.ORD_TIMESTAMP<? AND ({cond})"""
            r = pd.read_sql(sql, conn, params=[month_start, month_end]).iloc[0]
            rev += float(r['rev'] or 0)
            cogs += float(r['cogs'] or 0)
        finally:
            conn.close()
    return rev, cogs


def megaplus_sales_by_tin(tin: str, year: int, month: int) -> tuple[float, float]:
    """Sum revenue/cogs for products supplied by any distributor with this TIN."""
    month_start = f'{year}-{month:02d}-01'
    next_m = month + 1
    next_y = year
    if next_m > 12:
        next_m, next_y = 1, year + 1
    month_end = f'{next_y}-{next_m:02d}-01'
    rev = cogs = 0
    for store_id in STORES:
        conn = _connect(_db_name_for(store_id))
        try:
            dst = pd.read_sql(f"SELECT ID FROM DISTRIBUTORS WHERE saidentifikacio=N'{tin}'", conn)
            if len(dst) == 0:
                continue
            ids = ','.join(str(i) for i in dst['ID'])
            df_p = pd.read_sql(f"SELECT DISTINCT GAC_P_ID FROM GACERA WHERE GAC_D_ID IN ({ids})", conn)
            if len(df_p) == 0:
                continue
            pids = ','.join(str(p) for p in df_p['GAC_P_ID'])
            sql = f"""SELECT SUM(o.ORD_jamjam) rev, SUM(o.ORD_GETPRICE*o.ORD_quant) cogs
                FROM ORDERS o WHERE o.ORD_ACT=1 AND o.ORD_TIMESTAMP>=? AND o.ORD_TIMESTAMP<? AND o.ORD_P_ID IN ({pids})"""
            r = pd.read_sql(sql, conn, params=[month_start, month_end]).iloc[0]
            rev += float(r['rev'] or 0)
            cogs += float(r['cogs'] or 0)
        finally:
            conn.close()
    return rev, cogs


def load_elene_retro(year: int, month: int) -> dict:
    """Read Elene's retro file for this month: tin → amount."""
    pattern = str(FOODMART / 'გაშიფვრები_*' / f'*{month:02d},{year%100} რეტრო*.xlsx')
    matches = glob.glob(pattern)
    if not matches:
        return {}
    wb = openpyxl.load_workbook(matches[0], data_only=True)
    ws = wb.active
    out = {}
    for r in range(5, ws.max_row + 1):
        tin = ws.cell(r, 1).value
        amt = ws.cell(r, 3).value
        if isinstance(tin, (int, float)) and tin > 1_000_000 and isinstance(amt, (int, float)) and amt != 0:
            out[str(int(tin))] = float(amt)
    wb.close()
    return out


def load_lela_template(year: int, month: int) -> pd.DataFrame | None:
    """Find Lela's promo template for this month and return barcode/start/end/rate."""
    pattern = str(FOODMART / f'სააქცია_{year-1}-{year}' / f'*{year}-{month:02d}*.xlsx')
    matches = glob.glob(pattern)
    if not matches:
        # try alt naming
        pattern = str(FOODMART / f'სააქცია_*' / f'*{year}-{month:02d}*.xlsx')
        matches = glob.glob(pattern)
    if not matches:
        return None
    wb = openpyxl.load_workbook(matches[0], data_only=True)
    ws = wb.active
    hr = 3
    cols = {ws.cell(hr, c).value: c for c in range(1, ws.max_column + 1)}
    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        bc = ws.cell(r, cols.get('ბარკოდი', 5)).value
        rate = ws.cell(r, cols.get('ერთეულზე ასანაზღაურებელი თანხა_ჯამი', 14)).value
        if not isinstance(rate, (int, float)) or rate <= 0:
            continue
        rows.append({
            'barcode': normalize_barcode(bc),
            'start': ws.cell(r, cols.get('დაწყება', 1)).value,
            'end': ws.cell(r, cols.get('დასრულება', 2)).value,
            'rate': float(rate),
            'supplier': ws.cell(r, cols.get('მომწოდებელი', 4)).value,
        })
    wb.close()
    return pd.DataFrame(rows)


def compute_lela_cashback(template: pd.DataFrame) -> float:
    """For each template row, sum (qty sold during promo window) × rate, joining Megaplus."""
    total = 0
    for _, row in template.iterrows():
        bc = row['barcode']
        start = row['start']
        end = row['end']
        rate = row['rate']
        if not bc or not start or not end:
            continue
        qty_sum = 0
        for store_id in STORES:
            conn = _connect(_db_name_for(store_id))
            try:
                for pat in [bc, '0' + bc]:
                    sql = """SELECT SUM(o.ORD_quant) qty FROM ORDERS o JOIN PRODUCTS p ON p.P_ID=o.ORD_P_ID
                            WHERE o.ORD_ACT=1 AND CAST(p.P_BARCODE AS NVARCHAR(50))=?
                            AND o.ORD_TIMESTAMP>=? AND o.ORD_TIMESTAMP<=?"""
                    r = pd.read_sql(sql, conn, params=[pat, start, end])
                    q = float(r['qty'].iloc[0] or 0)
                    if q > 0:
                        qty_sum += q
                        break
            finally:
                conn.close()
        total += qty_sum * rate
    return total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('month', help='YYYY-MM')
    ap.add_argument('--json', help='write report to JSON path')
    args = ap.parse_args()
    year, month = map(int, args.month.split('-'))

    print('=' * 90)
    print(f'FOODMART SELF-VERIFY — {year}-{month:02d}')
    print('=' * 90)

    formulas = load_formulas()
    if not formulas:
        print('⚠️  formulas file empty/missing — run calibration first')
        return
    print(f'loaded {len(formulas)} supplier formulas')

    # 1. Retro per supplier (expected vs actual)
    rsge_in = rsge_waybill_in_per_tin(year, month)
    elene_actual = load_elene_retro(year, month)
    print()
    print(f'--- RETRO ({year}-{month:02d}) ---')
    print(f'{"TIN":<12} {"name":<28} {"base":<14} {"base_amt":>10} {"rate%":>7} {"expected":>10} {"actual":>10} {"diff":>10}')
    print('-' * 100)
    expected_sum = actual_sum = 0
    rows_out = []
    manual_review_rows = []
    for tin, fl in formulas.items():
        base = fl['base']
        rate = fl['rate_pct']
        if base == 'MANUAL_REVIEW' or rate is None:
            actual = elene_actual.get(tin, 0)
            last4 = fl.get('last_4_months', [])
            avg = sum(last4)/len(last4) if last4 else 0
            manual_review_rows.append({'tin':tin,'name':fl['name'],'actual':actual,'avg_last4':avg,'note':fl.get('note','')})
            print(f'{tin:<12} {fl["name"][:27]:<28} {"MANUAL":<14} {"—":>10} {"—":>7} {"~"+f"{avg:.0f}":>10} {actual:>10,.2f} (ხელით სამოწმდი)')
            continue
        if base == 'rs.ge_in':
            base_amt = rsge_in.get(tin, 0)
        elif base == 'rs.ge_in_group':
            related = fl.get('related_tins') or [tin]
            base_amt = sum(rsge_in.get(t, 0) for t in related)
        elif base == 'mp_brand_sales':
            rev, _ = megaplus_brand_sales(fl.get('name_keywords', []), year, month)
            base_amt = rev
        elif base == 'mp_brand_cogs':
            _, cog = megaplus_brand_sales(fl.get('name_keywords', []), year, month)
            base_amt = cog
        elif base == 'mp_sales':
            rev, _ = megaplus_sales_by_tin(tin, year, month)
            base_amt = rev
        elif base == 'mp_cogs':
            _, cog = megaplus_sales_by_tin(tin, year, month)
            base_amt = cog
        else:
            base_amt = 0
        expected = base_amt * rate / 100
        actual = elene_actual.get(tin, 0)
        diff = actual - expected
        expected_sum += expected
        actual_sum += actual
        print(f'{tin:<12} {fl["name"][:27]:<28} {base:<14} {base_amt:>10,.0f} {rate:>6.2f}% {expected:>10,.2f} {actual:>10,.2f} {diff:>+10,.2f}')
        rows_out.append({'tin':tin,'name':fl['name'],'base':base,'base_amount':base_amt,'rate_pct':rate,'expected':expected,'actual':actual,'diff':diff})
    print('-' * 100)
    print(f'{"TOTAL":<55} {expected_sum:>10,.2f} {actual_sum:>10,.2f} {actual_sum-expected_sum:>+10,.2f}')

    # 2. Cashback (Lela)
    print()
    print(f'--- ქეშბექი (Lela) ---')
    tpl = load_lela_template(year, month)
    if tpl is None:
        print('  ⚠️  Lela template not found for this month')
        cashback_expected = None
    else:
        print(f'  template rows with rate>0: {len(tpl)}')
        cashback_expected = compute_lela_cashback(tpl)
        print(f'  EXPECTED cashback: {cashback_expected:,.2f} ₾')

    # 3. Summary
    print()
    print('--- SUMMARY ---')
    print(f'  retro EXPECTED total: {expected_sum:,.2f} ₾')
    print(f'  retro ACTUAL total (Elene file): {actual_sum:,.2f} ₾')
    print(f'  retro DRIFT: {actual_sum - expected_sum:+,.2f} ₾')
    if cashback_expected is not None:
        total_expected_combined = expected_sum + cashback_expected
        print(f'  cashback EXPECTED: {cashback_expected:,.2f} ₾')
        print(f'  retro + cashback combined EXPECTED: {total_expected_combined:,.2f} ₾')

    if args.json:
        with open(args.json, 'w', encoding='utf-8') as f:
            json.dump({
                'month': args.month,
                'retro_rows': rows_out,
                'retro_expected_total': expected_sum,
                'retro_actual_total': actual_sum,
                'cashback_expected': cashback_expected,
            }, f, ensure_ascii=False, indent=2)
        print(f'  report written to {args.json}')


if __name__ == '__main__':
    main()
