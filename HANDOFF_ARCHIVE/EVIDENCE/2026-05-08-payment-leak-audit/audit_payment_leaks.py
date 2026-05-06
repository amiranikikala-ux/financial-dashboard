"""Audit v2: find REAL leaks — bank lines pipeline didn't count.

For each supplier:
  pipeline_counted = data["supplier_payment_lines"][tax_id]  # (date, amount, source)
  raw_bank_matched = lines I matched (by tax_id or strict name) for that supplier
  leak = raw_bank_matched − pipeline_counted   (by date+amount key)
"""
import os
import re
import json
from datetime import datetime
import openpyxl


def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S","%Y-%m-%d","%d.%m.%Y","%d/%m/%Y"):
            try: return datetime.strptime(s, fmt)
            except: pass
    return None


def to_float(v):
    if v is None or v == "": return None
    try: return float(v)
    except: return None


def normalize_name(n):
    if not n: return ""
    s = str(n).strip().strip("'\"")
    s = re.sub(r"^(შპს|სს|ი\.მ\.|ინდ\.მეწ\.|ააიპ|სსიპ|ააფ|საფ)\s+", "", s)
    s = s.strip("'\"").strip()
    return re.sub(r"\s+", " ", s).lower()


def parse_tbc_year(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            d = parse_date(row[0]) if len(row) > 0 else None
            if not d: continue
            paid_out = to_float(row[3]) if len(row) > 3 else None
            if paid_out is None or paid_out <= 0: continue
            partner_name = str(row[10]).strip() if len(row) > 10 and row[10] else ""
            partner_tax = str(row[11]).strip() if len(row) > 11 and row[11] else ""
            descr = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            out.append({"bank":"TBC","file":os.path.basename(path),"date":d.strftime("%Y-%m-%d"),
                        "amount":paid_out,"partner_name":partner_name,"partner_tax":partner_tax,"descr":descr})
    wb.close()
    return out


def parse_tbc_summary(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        if ws.title != "transactions_history": continue
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            d = parse_date(row[0]) if len(row) > 0 else None
            if not d: continue
            amt = to_float(row[3]) if len(row) > 3 else None
            if amt is None or amt >= 0: continue
            paid_out = -amt
            partner_name = str(row[11]).strip() if len(row) > 11 and row[11] else ""
            partner_tax = str(row[12]).strip() if len(row) > 12 and row[12] else ""
            descr = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            out.append({"bank":"TBC","file":os.path.basename(path),"date":d.strftime("%Y-%m-%d"),
                        "amount":paid_out,"partner_name":partner_name,"partner_tax":partner_tax,"descr":descr})
    wb.close()
    return out


def parse_bog(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        if ws.title != "Statement of Account": continue
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 9: continue
            d = parse_date(row[0]) if len(row) > 0 else None
            if not d: continue
            debit = to_float(row[5]) if len(row) > 5 else None
            if debit is None or debit <= 0: continue
            tx_type = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            if tx_type == "COM": continue
            descr = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            recipient_name = str(row[16]).strip() if len(row) > 16 and row[16] else ""
            recipient_tax = str(row[17]).strip() if len(row) > 17 and row[17] else ""
            if recipient_tax == "400333858": continue
            out.append({"bank":"BOG","file":os.path.basename(path),"date":d.strftime("%Y-%m-%d"),
                        "amount":debit,"partner_name":recipient_name,"partner_tax":recipient_tax,"descr":descr})
    wb.close()
    return out


def main():
    with open("rs-dashboard/public/data.json", encoding="utf-8") as f:
        data = json.load(f)

    suppliers_raw = data.get("suppliers", [])
    suppliers = []
    for s in suppliers_raw:
        m = re.match(r"\((\d+)\)\s*(.+)", s.get("ორგანიზაცია") or "")
        if not m: continue
        tax = m.group(1)
        suppliers.append({
            "tax_id": tax,
            "name_raw": m.group(2).strip(),
            "core": normalize_name(m.group(2)),
            "total_effective": s.get("total_effective") or 0,
            "system_paid": s.get("total_paid") or 0,
            "bank_paid": s.get("bank_paid") or 0,
            "manual_paid": s.get("manual_paid") or 0,
            "total_debt": s.get("total_debt") or 0,
        })
    by_tax = {s["tax_id"]: s for s in suppliers}
    by_core = {}
    for s in suppliers:
        if len(s["core"]) >= 4:
            by_core.setdefault(s["core"], []).append(s)

    pipeline_lines = data.get("supplier_payment_lines", {})

    # Build pipeline-counted set per supplier as (date, round(amount,2))
    pipeline_set = {}
    for tax, lines in pipeline_lines.items():
        s = set()
        for ln in lines:
            d = (ln.get("date") or "")[:10]
            a = round(float(ln.get("amount") or 0), 2)
            s.add((d, a))
        pipeline_set[tax] = s

    # Parse raw bank
    txs = []
    tbc_root = r"Financial_Analysis/თბს ბანკი ამონაწერი"
    bog_root = r"Financial_Analysis/ბოგ ბანკი ამონაწერი"
    for f in sorted(os.listdir(tbc_root)):
        path = os.path.join(tbc_root, f)
        if not f.endswith(".xlsx"): continue
        if f == "03.2026.xlsx":
            txs += parse_tbc_summary(path)
        else:
            txs += parse_tbc_year(path)
    for f in sorted(os.listdir(bog_root)):
        path = os.path.join(bog_root, f)
        if not f.endswith(".xlsx"): continue
        txs += parse_bog(path)

    # For each tx, attempt to attribute to a supplier
    leaks_per_supplier = {}
    matched_in_pipeline = 0
    leak_count = 0

    for tx in txs:
        tax = tx["partner_tax"]
        pname = normalize_name(tx["partner_name"])
        pname = re.sub(r",?\s*\d{8,12}\s*$", "", pname).strip()

        target_tax = None
        match_kind = None
        if tax and tax in by_tax:
            target_tax = tax
            match_kind = "by_tax"
        elif pname and pname in by_core:
            cands = by_core[pname]
            if len(cands) == 1:
                target_tax = cands[0]["tax_id"]
                match_kind = "by_name"

        if not target_tax: continue

        d = tx["date"][:10]
        a = round(tx["amount"], 2)
        key = (d, a)

        if key in pipeline_set.get(target_tax, set()):
            matched_in_pipeline += 1
        else:
            leaks_per_supplier.setdefault(target_tax, []).append({**tx, "match_kind": match_kind})
            leak_count += 1

    print(f"Total raw outflow tx: {len(txs)}")
    print(f"Attributed and counted by pipeline: {matched_in_pipeline}")
    print(f"Attributed but MISSED by pipeline: {leak_count}")
    print()

    rows = []
    for s in suppliers:
        leaks = leaks_per_supplier.get(s["tax_id"], [])
        if not leaks: continue
        leak_sum = sum(l["amount"] for l in leaks)
        leak_pre2023 = sum(l["amount"] for l in leaks if l["date"] < "2023-01-01")
        leak_2023plus = leak_sum - leak_pre2023
        leak_by_name = sum(l["amount"] for l in leaks if l["match_kind"] == "by_name")
        rows.append({
            "tax_id": s["tax_id"],
            "name": s["core"][:40],
            "effective": s["total_effective"],
            "system_paid": s["system_paid"],
            "leak_total": leak_sum,
            "leak_n": len(leaks),
            "leak_pre2023": leak_pre2023,
            "leak_2023plus": leak_2023plus,
            "leak_by_name": leak_by_name,
            "dashboard_debt": s["total_debt"],
            "real_debt": max(0, s["total_debt"] - leak_sum),
        })

    rows.sort(key=lambda r: -r["leak_total"])

    print(f"Suppliers with REAL leaks (uncounted by pipeline): {len(rows)}")
    print()
    print(f"{'tax_id':<12} {'name':<35} {'eff':>10} {'sys_paid':>10} {'leak':>10}({'n':>3}) {'pre2023':>9} {'2023+':>9} {'dash_debt':>10} {'real_debt':>10}")
    print("-" * 130)
    for r in rows[:60]:
        print(
            f"{r['tax_id']:<12} {r['name']:<35} "
            f"{r['effective']:>10,.0f} {r['system_paid']:>10,.0f} "
            f"{r['leak_total']:>10,.0f}({r['leak_n']:>3}) "
            f"{r['leak_pre2023']:>9,.0f} {r['leak_2023plus']:>9,.0f} "
            f"{r['dashboard_debt']:>10,.0f} {r['real_debt']:>10,.0f}"
        )

    sum_leak = sum(r["leak_total"] for r in rows)
    sum_pre = sum(r["leak_pre2023"] for r in rows)
    sum_post = sum(r["leak_2023plus"] for r in rows)
    sum_dash = sum(r["dashboard_debt"] for r in rows)
    sum_real = sum(r["real_debt"] for r in rows)
    print()
    print("=== TOTALS ===")
    print(f"Affected suppliers:          {len(rows)}")
    print(f"Leaked sum (all):            {sum_leak:>15,.2f} ₾")
    print(f"  pre-2023:                  {sum_pre:>15,.2f} ₾")
    print(f"  2023+:                     {sum_post:>15,.2f} ₾")
    print(f"Dashboard debt sum:          {sum_dash:>15,.2f} ₾")
    print(f"Estimated real debt sum:     {sum_real:>15,.2f} ₾")
    print(f"Reduction in apparent debt:  {sum_dash - sum_real:>15,.2f} ₾")

    with open("_audit_leaks_v2.json", "w", encoding="utf-8") as fout:
        json.dump({"leaks_per_supplier": leaks_per_supplier, "summary_rows": rows},
                  fout, ensure_ascii=False, indent=2)
    print("\nSaved → _audit_leaks_v2.json")


if __name__ == "__main__":
    main()
