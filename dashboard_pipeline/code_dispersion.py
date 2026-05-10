"""
Code dispersion detector — read-only utility.

Identifies MegaPlus PRODUCTS rows where the same physical product is
registered under multiple DIFFERENT codes (different P_IDs with different
barcodes). Symptom: one code holds heavy negative P_QUANT while one or more
sibling codes hold compensating positive P_QUANT — intake was recorded on
one code, sales on another.

Distinct from PRODUCTS_FRAGMENTATION (`HANDOFF_ARCHIVE/PREVIEWS/
PRODUCTS_FRAGMENTATION_2026-05-03.md`) which detects same-barcode-multiple-
P_IDs. Here we detect different-barcodes-same-physical-product.

Pipeline:
  1. SQL — anchors (P_QUANT < 0 AND ORDERS history) + positive-qty pool
  2. Prefilter — first-content-token-stem match (cheap O(N×k))
  3. AI classification (Haiku 4.5) — handles Georgian inflection, brand
     recognition, size normalization, variant disambiguation
  4. Math validation — total_in − total_out ≈ Σ qty_now (closure check)
  5. Excel review file — owner reviews verdicts and applies merges manually
     in MegaPlus UI; we never write to DB

Mirrors `orphan_resolver.py` shape: one Excel out, no DB writes, owner is
the final authority. NEVER writes back to MegaPlus.

Usage (PowerShell, parent venv):

    & "C:\\Users\\tengiz\\OneDrive\\Desktop\\AI აგენტი\\venv\\Scripts\\python.exe" \\
        -m dashboard_pipeline.code_dispersion --limit 200
"""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import anthropic
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from dashboard_pipeline.ai.config import load_ai_config
from dashboard_pipeline.megaplus_backup import _connect

logger = logging.getLogger(__name__)

STORES: dict[str, str] = {"1329": "დვაბზუ", "1301": "ოზურგეთი"}

CACHE_PATH = Path(__file__).resolve().parent.parent / ".code_dispersion_cache.json"
DESKTOP = Path(r"C:\Users\tengiz\OneDrive\Desktop")

# AI model + prompt
AI_MODEL = "claude-haiku-4-5-20251001"
AI_MAX_TOKENS = 1500
AI_TEMPERATURE = 0.0
AI_PARALLELISM = 6
AI_TIMEOUT_S = 60.0
AI_MAX_RETRIES = 3

SYSTEM_PROMPT = """\
You are a product-data analyst working with a Georgian retail store database.
Your job: decide which candidate products are the SAME physical product as an
anchor product, considering Georgian language morphology.

Rules for "same physical product":
- Same brand (ფეირი = Fairy; საირმე ≠ ნაბეღლავი ≠ ბორჯომი)
- Same flavor/variant (ლიმონი ≠ ფორთოხალი; Compact Blue ≠ Expert Red)
- Same per-unit size after normalization. Treat operator typos:
    * "0.450მლ" likely means 0.450ლ = 450მლ
    * "0.5ლ" = 500მლ; "21x450მლ" → unit = 450მლ; "(12ც) 0.150ლ" → 150მლ
- Same packaging type when distinguishable (cans ≠ glass ≠ PET)
- Different sizes (0.150ლ ≠ 0.250ლ) → NOT same product
- Inflected Georgian words (ლიმონი / ლიმონის / ლიმონს) → SAME stem
- Internal/short barcodes (4 digits, "1028", "4038") are MegaPlus-internal; do
  NOT use barcode equality to decide — judge by name + size only
- If anchor's name doesn't match candidate's brand+flavor+size, return NO

Return STRICT JSON only — no prose outside JSON. Schema:
{
  "anchor_p_id": <int>,
  "verdicts": [
    {"p_id": <int>, "same": true|false, "confidence": "high"|"medium"|"low",
     "reason_ka": "<short Georgian explanation>"},
    ...
  ]
}
"""


# ─────────────────────────── Normalization helpers ──────────────────────────

_token_re = re.compile(r"[\s/+\-_().,*\\\[\]]+")

# Generic prefixes that don't identify a brand on their own. The detector
# skips them when extracting first-content-stem.
_GENERIC_PREFIXES = {
    "მინ", "წყალი", "გაზ", "სასმელი", "შ", "მ", "შიდა", "მოხმარება",
    "სიგარეტი", "ერთჯერადი", "საღეჭი", "რეზინი", "ნამცხვარი", "ხორცი",
    "რძე", "ყველი", "მაკარონი", "შოკოლადი", "შოკოლადის",
    "ფქვილი", "ნაყინი", "ფაფა", "ფაფი",
}


def normalize(name: str) -> str:
    n = unicodedata.normalize("NFC", str(name or "")).strip().lower()
    return re.sub(r"\s+", " ", n)


def first_content_stem(name: str, stem_len: int = 4) -> str:
    """First non-generic non-numeric content token's prefix.

    Used as cheap O(1) prefilter key. Length-4 prefix tolerates Georgian
    case inflection (ლიმონი/ლიმონის/ლიმონს all share stem "ლიმო").
    """
    for tok in _token_re.split(normalize(name)):
        if not tok or len(tok) < 3:
            continue
        if tok in _GENERIC_PREFIXES:
            continue
        if re.match(r"^\d", tok):
            continue
        return tok[:stem_len]
    return ""


# ─────────────────────────────── Data fetch ────────────────────────────────

@dataclass
class Product:
    p_id: int
    barcode: str
    name: str
    qty: float
    supplier_uuid: str
    qty_in: float = 0.0
    qty_out: float = 0.0
    last_intake_g_time: Optional[int] = None  # raw yymmddhhxx


def _fetch_products(cur) -> list[Product]:
    cur.execute(
        """
        SELECT P_ID, LTRIM(RTRIM(P_BARCODE)) AS bc, P_NAME, P_QUANT,
               P_DAFAULTSUPPLIER
        FROM PRODUCTS
        WHERE P_NAME IS NOT NULL AND LTRIM(RTRIM(P_NAME)) <> ''
        """
    )
    return [
        Product(
            p_id=int(r.P_ID),
            barcode=r.bc or "",
            name=r.P_NAME,
            qty=float(r.P_QUANT or 0),
            supplier_uuid=str(r.P_DAFAULTSUPPLIER) if r.P_DAFAULTSUPPLIER else "",
        )
        for r in cur.fetchall()
    ]


_IN_CHUNK = 1500  # SQL Server cap is ~2100 params per statement


def _fetch_flow_totals(
    cur, p_ids: list[int]
) -> tuple[dict[int, float], dict[int, float], dict[int, int]]:
    """Return (qty_in_by_pid, qty_out_by_pid, last_intake_g_time_by_pid).

    `last_intake_g_time_by_pid` returns raw G_TIME bigint (yymmddhhxx). Use
    `parse_g_time` from waybill_reconciliation to decode to date.

    Chunks IN-clause to stay under SQL Server's 2100-parameter limit.
    """
    qty_in: dict[int, float] = {}
    qty_out: dict[int, float] = {}
    last_intake: dict[int, int] = {}
    for i in range(0, len(p_ids), _IN_CHUNK):
        chunk = p_ids[i:i + _IN_CHUNK]
        if not chunk:
            continue
        placeholders = ",".join(["?"] * len(chunk))
        cur.execute(
            f"""
            SELECT G_P_ID, SUM(G_QUANT) AS s, MAX(G_TIME) AS lt FROM GET
            WHERE G_ACT = 1 AND G_P_ID IN ({placeholders})
            GROUP BY G_P_ID
            """,
            *chunk,
        )
        for r in cur.fetchall():
            qty_in[int(r.G_P_ID)] = float(r.s or 0)
            if r.lt is not None:
                last_intake[int(r.G_P_ID)] = int(r.lt)
        cur.execute(
            f"""
            SELECT ORD_P_ID, SUM(ORD_QUANT) AS s FROM ORDERS
            WHERE ORD_ACT = 1 AND ORD_P_ID IN ({placeholders})
            GROUP BY ORD_P_ID
            """,
            *chunk,
        )
        for r in cur.fetchall():
            qty_out[int(r.ORD_P_ID)] = float(r.s or 0)
    return qty_in, qty_out, last_intake


# ─────────────────────────────── Prefilter ─────────────────────────────────

def build_candidates(
    anchor: Product,
    pos_pool: list[Product],
    pos_by_stem: dict[str, list[Product]],
    top_k: int = 12,
) -> list[Product]:
    """Stem-prefix match + Jaccard token-overlap ranking → top-k candidates."""
    stem = first_content_stem(anchor.name)
    if not stem:
        return []
    pool = pos_by_stem.get(stem, [])
    a_tokens = set(t for t in _token_re.split(normalize(anchor.name)) if t)
    scored: list[tuple[int, Product]] = []
    for p in pool:
        if p.p_id == anchor.p_id:
            continue
        p_tokens = set(t for t in _token_re.split(normalize(p.name)) if t)
        overlap = len(a_tokens & p_tokens)
        scored.append((overlap, p))
    scored.sort(key=lambda x: -x[0])
    return [p for _, p in scored[:top_k]]


# ─────────────────────────────── AI client ─────────────────────────────────

@dataclass
class AIVerdict:
    p_id: int
    same: bool
    confidence: str
    reason_ka: str


@dataclass
class AICall:
    anchor_pid: int
    verdicts: list[AIVerdict]
    input_tokens: int
    output_tokens: int
    error: Optional[str] = None


def _strip_code_fence(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.MULTILINE).strip()
    return text


def call_ai_classify(
    client: anthropic.Anthropic, anchor: Product, candidates: list[Product]
) -> AICall:
    payload = {
        "anchor": {"p_id": anchor.p_id, "name": anchor.name, "barcode": anchor.barcode},
        "candidates": [
            {"p_id": c.p_id, "name": c.name, "barcode": c.barcode}
            for c in candidates
        ],
    }
    last_err: Optional[Exception] = None
    for attempt in range(AI_MAX_RETRIES):
        try:
            resp = client.messages.create(
                model=AI_MODEL,
                max_tokens=AI_MAX_TOKENS,
                temperature=AI_TEMPERATURE,
                system=SYSTEM_PROMPT,
                messages=[
                    {"role": "user", "content": json.dumps(payload, ensure_ascii=False)}
                ],
            )
            text = "".join(b.text for b in resp.content if hasattr(b, "text"))
            text = _strip_code_fence(text)
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError as e:
                return AICall(
                    anchor_pid=anchor.p_id, verdicts=[],
                    input_tokens=resp.usage.input_tokens,
                    output_tokens=resp.usage.output_tokens,
                    error=f"json parse: {e}; raw[:200]={text[:200]!r}",
                )
            verdicts = [
                AIVerdict(
                    p_id=int(v.get("p_id", 0)),
                    same=bool(v.get("same", False)),
                    confidence=str(v.get("confidence", "low")),
                    reason_ka=str(v.get("reason_ka", "")),
                )
                for v in parsed.get("verdicts", [])
            ]
            return AICall(
                anchor_pid=anchor.p_id, verdicts=verdicts,
                input_tokens=resp.usage.input_tokens,
                output_tokens=resp.usage.output_tokens,
            )
        except (anthropic.APIConnectionError, anthropic.APITimeoutError, anthropic.RateLimitError) as e:
            last_err = e
            time.sleep(2 ** attempt)
            continue
        except Exception as e:
            last_err = e
            break
    return AICall(
        anchor_pid=anchor.p_id, verdicts=[], input_tokens=0, output_tokens=0,
        error=f"{type(last_err).__name__}: {last_err}",
    )


# ─────────────────────────────── Cache ─────────────────────────────────────

def _cache_key(anchor: Product, candidates: list[Product]) -> str:
    """Deterministic key — same anchor+same candidate list → same cache hit."""
    return f"{anchor.p_id}|" + ",".join(str(c.p_id) for c in sorted(candidates, key=lambda c: c.p_id))


def load_cache() -> dict[str, dict]:
    if CACHE_PATH.exists():
        try:
            return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {}
    return {}


def save_cache(cache: dict[str, dict]) -> None:
    CACHE_PATH.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ─────────────────────────────── Group assembly ────────────────────────────

@dataclass
class DispersionGroup:
    store_id: str
    store_label: str
    anchor: Product
    siblings: list[Product]
    rejected: list[tuple[Product, AIVerdict]]
    confidence: str           # "high" / "medium" / "low" / "no_match"
    family_qty_now: float
    family_qty_in: float
    family_qty_out: float
    balance_check: float      # qty_in - qty_out (should ≈ family_qty_now if closed)
    balance_consistent: bool
    economic_score: float     # 0..1; 1 = perfect closure
    ai_input_tokens: int
    ai_output_tokens: int
    ai_error: Optional[str] = None
    canonical_pid: Optional[int] = None
    canonical_reason_ka: str = ""
    actions: list[dict] = field(default_factory=list)


def _aggregate_confidence(siblings_verdicts: list[AIVerdict]) -> str:
    if not siblings_verdicts:
        return "no_match"
    confidences = [v.confidence.lower() for v in siblings_verdicts]
    if any(c == "low" for c in confidences):
        return "low"
    if any(c == "medium" for c in confidences):
        return "medium"
    return "high"


def assemble_group(
    store_id: str,
    store_label: str,
    anchor: Product,
    candidates: list[Product],
    ai_call: AICall,
) -> DispersionGroup:
    cand_by_pid = {c.p_id: c for c in candidates}
    siblings: list[Product] = []
    sibling_verdicts: list[AIVerdict] = []
    rejected: list[tuple[Product, AIVerdict]] = []
    for v in ai_call.verdicts:
        c = cand_by_pid.get(v.p_id)
        if c is None:
            continue
        if v.same:
            siblings.append(c)
            sibling_verdicts.append(v)
        else:
            rejected.append((c, v))

    family = [anchor] + siblings
    family_qty_now = sum(p.qty for p in family)
    family_qty_in = sum(p.qty_in for p in family)
    family_qty_out = sum(p.qty_out for p in family)
    balance_check = family_qty_in - family_qty_out
    abs_diff = abs(balance_check - family_qty_now)
    flow_basis = max(family_qty_in, family_qty_out, abs(family_qty_now), 1.0)
    economic_score = max(0.0, 1.0 - abs_diff / flow_basis)
    balance_consistent = abs_diff < 1.0

    grp = DispersionGroup(
        store_id=store_id,
        store_label=store_label,
        anchor=anchor,
        siblings=siblings,
        rejected=rejected,
        confidence=_aggregate_confidence(sibling_verdicts),
        family_qty_now=family_qty_now,
        family_qty_in=family_qty_in,
        family_qty_out=family_qty_out,
        balance_check=balance_check,
        balance_consistent=balance_consistent,
        economic_score=economic_score,
        ai_input_tokens=ai_call.input_tokens,
        ai_output_tokens=ai_call.output_tokens,
        ai_error=ai_call.error,
    )
    if siblings:
        recommend_canonical(grp)
    return grp


def _looks_like_real_ean(barcode: str) -> bool:
    """EAN-like barcode = 8+ digits, all numeric. Internal/short codes
    (`4038`, `1062`, `2601...`) get scored lower."""
    bc = (barcode or "").strip()
    if not bc.isdigit():
        return False
    return len(bc) >= 8


def recommend_canonical(grp: DispersionGroup) -> None:
    """Pick the P_ID to keep + emit explicit transfer actions.

    Ranking key (winner = lowest tuple):
      1. Most-recent intake date (G_TIME) — the code currently being scanned
         at intake is what cashiers should also scan at sale
      2. Real EAN barcode preferred over internal short code
      3. Highest total intake quantity (longest-established code)
      4. Lowest P_ID (oldest, most-established)

    Sets `grp.canonical_pid`, `grp.canonical_reason_ka`, `grp.actions`.
    """
    family = [grp.anchor] + grp.siblings

    def rank_key(p: Product) -> tuple:
        return (
            -(p.last_intake_g_time or 0),   # most recent first (negate for asc sort)
            0 if _looks_like_real_ean(p.barcode) else 1,
            -p.qty_in,                       # most intake first
            p.p_id,
        )

    family_sorted = sorted(family, key=rank_key)
    canonical = family_sorted[0]
    losers = family_sorted[1:]

    reasons = []
    if canonical.last_intake_g_time:
        try:
            from dashboard_pipeline.waybill_reconciliation import parse_g_time
            d = parse_g_time(canonical.last_intake_g_time)
            if d is not None:
                reasons.append(f"ბოლო შემოსვლა {d:%Y-%m-%d}")
        except Exception:
            pass
    if _looks_like_real_ean(canonical.barcode):
        reasons.append(f"ნამდვილი ბარკოდი ({canonical.barcode})")
    if canonical.qty_in > 0:
        reasons.append(f"ჯამური შემოსვლა {int(canonical.qty_in)} ცალი")
    grp.canonical_pid = canonical.p_id
    grp.canonical_reason_ka = "; ".join(reasons) if reasons else f"P_ID {canonical.p_id} ყველაზე ძველი/აქტიური"

    actions: list[dict] = []
    for loser in losers:
        if loser.qty == 0:
            actions.append({
                "from_pid": loser.p_id,
                "from_barcode": loser.barcode,
                "from_name": loser.name,
                "to_pid": canonical.p_id,
                "to_barcode": canonical.barcode,
                "to_name": canonical.name,
                "qty_to_transfer": 0.0,
                "instruction_ka": (
                    f"კოდი {loser.p_id} (ბარკოდი {loser.barcode or '—'}) — "
                    f"ნაშთი 0, უბრალოდ წაშალე ან ჩამოარეგისტრიე (P_ACTIVE = 0)"
                ),
            })
            continue
        actions.append({
            "from_pid": loser.p_id,
            "from_barcode": loser.barcode,
            "from_name": loser.name,
            "to_pid": canonical.p_id,
            "to_barcode": canonical.barcode,
            "to_name": canonical.name,
            "qty_to_transfer": loser.qty,
            "instruction_ka": (
                f"გადაიტანე {loser.qty:+g} ცალი კოდიდან {loser.p_id} "
                f"(ბარკოდი {loser.barcode or '—'}) კოდზე {canonical.p_id} "
                f"(ბარკოდი {canonical.barcode or '—'}). მერე ჩამოარეგისტრიე {loser.p_id}."
            ),
        })
    grp.actions = actions


# ─────────────────────────────── Excel writer ──────────────────────────────

_BOLD = Font(bold=True)
_FILL_GREEN = PatternFill("solid", fgColor="C8E6C9")
_FILL_YELLOW = PatternFill("solid", fgColor="FFF59D")
_FILL_RED = PatternFill("solid", fgColor="FFCDD2")
_FILL_BLUE = PatternFill("solid", fgColor="BBDEFB")


def write_excel(groups: list[DispersionGroup], out_path: Path) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: ქმედებები — action-oriented prescription, owner reads first ──
    ws_act = wb.active
    ws_act.title = "ქმედებები"
    act_headers = [
        "მაღაზია", "გატეხილი პროდუქტი (სახელი)", "გატეხილის ნაშთი",
        "AI-ს ნდობა", "ბალანსი იხურება?",
        "მთავარი კოდი (დარჩება)", "მთავარის ბარკოდი", "რატომ ეს კოდი",
        "კოდი წასაშლელი", "წასაშლელის ბარკოდი", "წასაშლელის ნაშთი",
        "ცხადი ინსტრუქცია",
        "შესრულდა? (✓ ან თარიღი)",
    ]
    ws_act.append(act_headers)
    for c in ws_act[1]:
        c.font = _BOLD

    sortable = sorted(groups, key=lambda g: (g.confidence != "high", g.anchor.qty))
    for g in sortable:
        if not g.siblings or not g.actions:
            continue
        conf_ka = {"high": "მაღალი", "medium": "საშუალო", "low": "დაბალი", "no_match": "—"}.get(g.confidence, g.confidence)
        family = [g.anchor] + g.siblings
        canonical = next((p for p in family if p.p_id == g.canonical_pid), g.anchor)
        bal_ka = "კი" if g.balance_consistent else "არა"
        for a in g.actions:
            ws_act.append([
                g.store_label, g.anchor.name, g.anchor.qty,
                conf_ka, bal_ka,
                canonical.p_id, canonical.barcode or "—", g.canonical_reason_ka,
                a["from_pid"], a["from_barcode"] or "—", a["qty_to_transfer"],
                a["instruction_ka"],
                "",
            ])
            if g.confidence == "high" and g.balance_consistent:
                for cell in ws_act[ws_act.max_row]:
                    cell.fill = _FILL_GREEN
            elif g.confidence in ("medium", "low") or not g.balance_consistent:
                for cell in ws_act[ws_act.max_row]:
                    cell.fill = _FILL_YELLOW

    # ── Sheet 2: ჯგუფები — high-level group summary ──
    ws_sum = wb.create_sheet("ჯგუფები")
    sum_headers = [
        "მაღაზია", "გატეხილი P_ID", "გატეხილის სახელი", "გატეხილის ნაშთი",
        "იგივე-კოდის რაოდენობა", "AI-ს ნდობა",
        "ჯგუფის ჯამი", "ჯამური შემოსვლა", "ჯამური გაყიდვა",
        "ბალანსის შემოწმება", "ბალანსი იხურება?", "მათემ. შეფასება",
        "გადაწყვეტილება (შენ ჩაწერე: გავაერთიანო / ცალკე / ვერ-მივხვდი)",
    ]
    ws_sum.append(sum_headers)
    for c in ws_sum[1]:
        c.font = _BOLD

    sortable = sorted(groups, key=lambda g: (g.confidence != "high", g.anchor.qty))
    for g in sortable:
        if not g.siblings:
            continue
        conf_ka = {"high": "მაღალი", "medium": "საშუალო", "low": "დაბალი", "no_match": "—"}.get(g.confidence, g.confidence)
        row = [
            g.store_label, g.anchor.p_id, g.anchor.name, g.anchor.qty,
            len(g.siblings), conf_ka,
            round(g.family_qty_now, 2), round(g.family_qty_in, 2), round(g.family_qty_out, 2),
            round(g.balance_check, 2), "კი" if g.balance_consistent else "არა",
            round(g.economic_score, 3),
            "",  # owner fills
        ]
        ws_sum.append(row)
        if g.confidence == "high" and g.balance_consistent:
            for cell in ws_sum[ws_sum.max_row]:
                cell.fill = _FILL_GREEN
        elif g.confidence in ("medium", "low"):
            for cell in ws_sum[ws_sum.max_row]:
                cell.fill = _FILL_YELLOW

    ws_det = wb.create_sheet("დეტალები")
    det_headers = [
        "მაღაზია", "გატეხილი P_ID", "AI-ს ნდობა", "როლი", "P_ID", "ბარკოდი",
        "სახელი", "ნაშთი", "შემოვიდა", "გაიყიდა", "მომწოდებლის UUID",
        "AI: იგივეა?", "AI-ს ნდობა (ხაზზე)", "AI-ს ახსნა",
    ]
    ws_det.append(det_headers)
    for c in ws_det[1]:
        c.font = _BOLD

    conf_map = {"high": "მაღალი", "medium": "საშუალო", "low": "დაბალი", "no_match": "—"}
    for g in sortable:
        if not g.siblings and not g.rejected:
            continue
        conf_ka = conf_map.get(g.confidence, g.confidence)
        # anchor row
        ws_det.append([
            g.store_label, g.anchor.p_id, conf_ka, "გატეხილი",
            g.anchor.p_id, g.anchor.barcode, g.anchor.name,
            g.anchor.qty, g.anchor.qty_in, g.anchor.qty_out, g.anchor.supplier_uuid,
            "", "", f"უარყოფითი ნაშთი = {g.anchor.qty}",
        ])
        for cell in ws_det[ws_det.max_row]:
            cell.fill = _FILL_BLUE
        # siblings (verdict YES rows)
        for s in g.siblings:
            ws_det.append([
                g.store_label, g.anchor.p_id, conf_ka, "იგივე-პროდუქტი",
                s.p_id, s.barcode, s.name,
                s.qty, s.qty_in, s.qty_out, s.supplier_uuid,
                "კი", conf_ka, "AI: იგივე პროდუქტია",
            ])
            for cell in ws_det[ws_det.max_row]:
                cell.fill = _FILL_GREEN
        # rejected (verdict NO rows — kept for transparency)
        for c, v in g.rejected:
            ws_det.append([
                g.store_label, g.anchor.p_id, conf_ka, "უარყოფილი",
                c.p_id, c.barcode, c.name,
                c.qty, c.qty_in, c.qty_out, c.supplier_uuid,
                "არა", conf_map.get(v.confidence, v.confidence), v.reason_ka,
            ])

    # Column widths
    for ws in (ws_sum, ws_det):
        for col_cells in ws.columns:
            ml = 0
            letter = col_cells[0].column_letter
            for cell in col_cells:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > ml:
                    ml = len(v)
            ws.column_dimensions[letter].width = min(70, max(10, ml + 2))

    wb.save(out_path)


# ─────────────────────────────── Per-store driver ──────────────────────────

def process_store(
    store_id: str,
    store_label: str,
    client: anthropic.Anthropic,
    cache: dict[str, dict],
    limit: Optional[int] = None,
    parallelism: int = AI_PARALLELISM,
) -> list[DispersionGroup]:
    db = f"MEGAPLUS_{store_id}"
    print(f"\n=== {store_label} ({db}) ===", flush=True)
    conn = _connect(db, autocommit=True)
    cur = conn.cursor()

    products = _fetch_products(cur)
    print(f"  PRODUCTS scanned: {len(products):,}", flush=True)

    anchors = sorted(
        [p for p in products if p.qty < 0], key=lambda p: p.qty
    )
    if limit is not None:
        anchors = anchors[:limit]

    pos_pool = [p for p in products if p.qty > 0]
    pos_by_stem: dict[str, list[Product]] = {}
    for p in pos_pool:
        s = first_content_stem(p.name)
        if s:
            pos_by_stem.setdefault(s, []).append(p)

    print(f"  anchors (qty<0, capped @ {limit}): {len(anchors):,}", flush=True)
    print(f"  positive-qty pool: {len(pos_pool):,}", flush=True)

    # ── Build candidates + qualify anchors that have ORDERS history ──
    pid_set = {a.p_id for a in anchors}
    for a in anchors:
        a.candidates = build_candidates(a, pos_pool, pos_by_stem, top_k=12)  # type: ignore[attr-defined]
        for c in a.candidates:  # type: ignore[attr-defined]
            pid_set.add(c.p_id)

    qty_in_by_pid, qty_out_by_pid, last_intake_by_pid = _fetch_flow_totals(cur, list(pid_set))
    for p in products:
        if p.p_id in qty_in_by_pid:
            p.qty_in = qty_in_by_pid[p.p_id]
        if p.p_id in qty_out_by_pid:
            p.qty_out = qty_out_by_pid[p.p_id]
        if p.p_id in last_intake_by_pid:
            p.last_intake_g_time = last_intake_by_pid[p.p_id]

    # Filter anchors: must have actual sales history (qty_out > 0) — otherwise
    # negative qty isn't from "code dispersion", it's likely a different bug
    qualified = [a for a in anchors if a.qty_out > 0 and a.candidates]  # type: ignore[attr-defined]
    print(f"  qualified anchors (have ORDERS + ≥1 candidate): {len(qualified):,}", flush=True)

    # ── AI classification (parallel, with cache) ──
    groups: list[DispersionGroup] = []
    new_cache_entries = 0

    def _classify(anchor: Product) -> DispersionGroup:
        nonlocal new_cache_entries
        candidates = anchor.candidates  # type: ignore[attr-defined]
        key = _cache_key(anchor, candidates)
        if key in cache:
            entry = cache[key]
            ai_call = AICall(
                anchor_pid=anchor.p_id,
                verdicts=[
                    AIVerdict(
                        p_id=int(v["p_id"]),
                        same=bool(v["same"]),
                        confidence=str(v["confidence"]),
                        reason_ka=str(v["reason_ka"]),
                    )
                    for v in entry.get("verdicts", [])
                ],
                input_tokens=int(entry.get("input_tokens", 0)),
                output_tokens=int(entry.get("output_tokens", 0)),
                error=entry.get("error"),
            )
        else:
            ai_call = call_ai_classify(client, anchor, candidates)
            cache[key] = {
                "verdicts": [
                    {"p_id": v.p_id, "same": v.same, "confidence": v.confidence,
                     "reason_ka": v.reason_ka}
                    for v in ai_call.verdicts
                ],
                "input_tokens": ai_call.input_tokens,
                "output_tokens": ai_call.output_tokens,
                "error": ai_call.error,
            }
            new_cache_entries += 1
        return assemble_group(store_id, store_label, anchor, candidates, ai_call)

    save_every = 25
    with ThreadPoolExecutor(max_workers=parallelism) as ex:
        futures = {ex.submit(_classify, a): a for a in qualified}
        for i, fut in enumerate(as_completed(futures), 1):
            try:
                grp = fut.result()
                groups.append(grp)
            except Exception as e:
                logger.exception("AI classification failed: %s", e)
            if i % 10 == 0:
                done = sum(1 for g in groups if g.siblings)
                print(f"    progress: {i}/{len(qualified)} | groups so far: {done}", flush=True)
            if new_cache_entries and new_cache_entries % save_every == 0:
                save_cache(cache)

    save_cache(cache)
    found = sum(1 for g in groups if g.siblings)
    closed = sum(1 for g in groups if g.siblings and g.balance_consistent)
    print(f"  groups with siblings: {found:,}  (math-closed: {closed:,})", flush=True)
    return groups


# ─────────────────────────────── Main ──────────────────────────────────────

def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="MegaPlus code-dispersion detector")
    parser.add_argument("--limit", type=int, default=None,
                        help="Per-store anchor cap (most-negative first). Default: no cap.")
    parser.add_argument("--store", choices=list(STORES.keys()) + ["all"], default="all")
    parser.add_argument("--parallelism", type=int, default=AI_PARALLELISM)
    parser.add_argument("--out", type=str, default=None,
                        help="Excel output path. Default: Desktop\\code_dispersion_review_<date>.xlsx")
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    cfg = load_ai_config()
    if not cfg.is_enabled:
        print("ANTHROPIC_API_KEY not configured — abort", file=sys.stderr)
        return 1
    print(f"AI ready: {cfg.redacted()}", flush=True)
    client = anthropic.Anthropic(api_key=cfg.api_key, timeout=AI_TIMEOUT_S)

    cache = load_cache()
    print(f"cache: {len(cache):,} entries at {CACHE_PATH}", flush=True)

    selected = list(STORES.items()) if args.store == "all" else [(args.store, STORES[args.store])]
    all_groups: list[DispersionGroup] = []
    for store_id, store_label in selected:
        groups = process_store(
            store_id, store_label, client, cache,
            limit=args.limit, parallelism=args.parallelism,
        )
        all_groups.extend(groups)

    out_path = (
        Path(args.out) if args.out
        else DESKTOP / f"code_dispersion_review_{datetime.now():%Y-%m-%d}.xlsx"
    )
    write_excel(all_groups, out_path)
    print(f"\n>>> Excel written: {out_path}", flush=True)

    total_in = sum(g.ai_input_tokens for g in all_groups)
    total_out = sum(g.ai_output_tokens for g in all_groups)
    cost_in = total_in / 1_000_000 * 1.00
    cost_out = total_out / 1_000_000 * 5.00
    print(f"\n=== AI usage ({len(all_groups):,} anchors) ===", flush=True)
    print(f"  input tokens:  {total_in:,}  →  ~${cost_in:.4f}", flush=True)
    print(f"  output tokens: {total_out:,}  →  ~${cost_out:.4f}", flush=True)
    print(f"  total: ~${cost_in + cost_out:.4f}", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
